"""
Generador de Plantilla Excel para Análisis Horizontal Financiero.

Uso:
    python generar_plantilla.py <output_path> [--empresa NOMBRE] [--tipo TIPO_EEFF]
        [--periodo-base PERIODO] [--periodo-comp PERIODO] [--moneda MONEDA] [--escala ESCALA]

Ejemplo:
    python generar_plantilla.py reporte.xlsx --empresa "Acme Corp" --tipo "P&L" \
        --periodo-base "2023" --periodo-comp "2024" --moneda "USD" --escala "Miles"

Si se omiten parámetros, se dejan como placeholders para llenar manualmente.
"""

import argparse
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter


# === ESTILOS ===
BLUE_DARK = "003366"
GRAY_LIGHT = "F2F2F2"
GRAY_MED = "D9D9D9"
GREEN_BG = "E8F5E9"
GREEN_TXT = "2E7D32"
RED_BG = "FFEBEE"
RED_TXT = "C62828"
BLUE_INPUT = "0000FF"
YELLOW_BG = "FFF9C4"
WHITE = "FFFFFF"

font_title = Font(name="Arial", size=16, bold=True, color=BLUE_DARK)
font_section = Font(name="Arial", size=12, bold=True, color=BLUE_DARK)
font_header = Font(name="Arial", size=10, bold=True, color=WHITE)
font_data = Font(name="Arial", size=10)
font_input = Font(name="Arial", size=10, color=BLUE_INPUT)
font_subtotal = Font(name="Arial", size=10, bold=True)
font_total = Font(name="Arial", size=10, bold=True, color=BLUE_DARK)
font_note = Font(name="Arial", size=9, italic=True, color="666666")
font_label = Font(name="Arial", size=10, bold=True)

fill_header = PatternFill("solid", fgColor=BLUE_DARK)
fill_gray_light = PatternFill("solid", fgColor=GRAY_LIGHT)
fill_gray_med = PatternFill("solid", fgColor=GRAY_MED)
fill_green = PatternFill("solid", fgColor=GREEN_BG)
fill_red = PatternFill("solid", fgColor=RED_BG)
fill_yellow = PatternFill("solid", fgColor=YELLOW_BG)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")
align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

border_thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
border_bottom_double = Border(
    bottom=Side(style="double", color=BLUE_DARK),
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, widths=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin


def create_resumen_ejecutivo(wb, params):
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    set_col_widths(ws, [20, 25, 20, 20, 20, 20])

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "INFORME DE ANÁLISIS HORIZONTAL"
    c.font = font_title
    c.alignment = align_center

    labels = [
        (3, "Empresa:", params.get("empresa", "[Nombre de la empresa]")),
        (4, "Estado Financiero:", params.get("tipo", "[Tipo de EEFF]")),
        (5, "Período Base:", params.get("periodo_base", "[Período base]")),
        (6, "Período Comparación:", params.get("periodo_comp", "[Período comparación]")),
        (7, "Moneda:", params.get("moneda", "[Moneda]")),
        (8, "Escala:", params.get("escala", "[Escala]")),
        (9, "Fecha de elaboración:", datetime.now().strftime("%Y-%m-%d")),
    ]
    for row, label, value in labels:
        ws.cell(row=row, column=1, value=label).font = font_label
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = font_input
        cell.border = border_thin

    row = 11
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="VEREDICTO GENERAL")
    c.font = font_section
    c.fill = fill_gray_light

    ws.cell(row=12, column=1, value="Semáforo:").font = font_label
    ws.cell(row=12, column=2, value="[🟢/🟡/🔴]").font = font_input
    ws.cell(row=13, column=1, value="Conclusión:").font = font_label
    ws.merge_cells("B13:F14")
    ws.cell(row=13, column=2, value="[Conclusión del análisis en 2-3 oraciones]").font = font_input
    ws["B13"].alignment = align_wrap

    row = 16
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="TOP 3 VARIACIONES MÁS SIGNIFICATIVAS")
    c.font = font_section
    c.fill = fill_gray_light

    for i, r in enumerate([17, 18, 19], 1):
        ws.cell(row=r, column=1, value=f"#{i}:").font = font_label
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=f"[Variación #{i}]").font = font_input

    ws.sheet_properties.pageSetUpPr = None
    return ws


def create_analisis_horizontal(wb, params):
    ws = wb.create_sheet("Análisis Horizontal")
    p_base = params.get("periodo_base", "Período Base")
    p_comp = params.get("periodo_comp", "Período Comp.")

    headers = ["Partida", p_base, p_comp, "Var. Absoluta", "Var. %", "Participación Base %", "Señal"]
    set_col_widths(ws, [35, 18, 18, 18, 12, 18, 15])
    write_header_row(ws, 1, headers)

    # Filas ejemplo para P&L
    sample_rows = [
        ("INGRESOS", None, None, True, False),
        ("  Ingresos operacionales", 0, 0, False, False),
        ("  Otros ingresos", 0, 0, False, False),
        ("TOTAL INGRESOS", None, None, False, True),
        ("", None, None, False, False),
        ("COSTOS", None, None, True, False),
        ("  Costo de ventas", 0, 0, False, False),
        ("  Costo de servicios", 0, 0, False, False),
        ("TOTAL COSTOS", None, None, False, True),
        ("", None, None, False, False),
        ("UTILIDAD BRUTA (Gross Profit)", None, None, False, True),
        ("", None, None, False, False),
        ("GASTOS OPERACIONALES", None, None, True, False),
        ("  Gastos de administración", 0, 0, False, False),
        ("  Gastos de ventas", 0, 0, False, False),
        ("  Depreciación y amortización", 0, 0, False, False),
        ("  Otros gastos operacionales", 0, 0, False, False),
        ("TOTAL GASTOS OPERACIONALES", None, None, False, True),
        ("", None, None, False, False),
        ("UTILIDAD OPERACIONAL (EBIT)", None, None, False, True),
        ("", None, None, False, False),
        ("RESULTADO FINANCIERO", None, None, True, False),
        ("  Ingresos financieros", 0, 0, False, False),
        ("  Gastos financieros", 0, 0, False, False),
        ("  Diferencia en cambio", 0, 0, False, False),
        ("TOTAL RESULTADO FINANCIERO", None, None, False, True),
        ("", None, None, False, False),
        ("RESULTADO ANTES DE IMPUESTOS (EBT)", None, None, False, True),
        ("  Impuesto de renta", 0, 0, False, False),
        ("UTILIDAD NETA (Net Income)", None, None, False, True),
    ]

    for i, (name, base, comp, is_section, is_total) in enumerate(sample_rows, 2):
        r = i
        ws.cell(row=r, column=1, value=name).font = font_subtotal if is_section else (font_total if is_total else font_data)
        if is_section:
            ws.cell(row=r, column=1).fill = fill_gray_light
        if is_total:
            ws.cell(row=r, column=1).fill = fill_gray_med
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = border_bottom_double

        if base is not None:
            ws.cell(row=r, column=2, value=base).font = font_input
            ws.cell(row=r, column=2).number_format = '#,##0'
            ws.cell(row=r, column=3, value=comp).font = font_input
            ws.cell(row=r, column=3).number_format = '#,##0'
            # Fórmulas
            ws.cell(row=r, column=4).value = f'=C{r}-B{r}'
            ws.cell(row=r, column=4).number_format = '#,##0;(#,##0);"-"'
            ws.cell(row=r, column=5).value = f'=IF(B{r}=0,IF(C{r}=0,0,"N/A"),D{r}/ABS(B{r})*100)'
            ws.cell(row=r, column=5).number_format = '0.0"%"'
            ws.cell(row=r, column=7).value = f'=IF(E{r}="N/A","🔵",IF(ABS(E{r})>15,IF(E{r}>0,"🟢 ↑↑","🔴 ↓↓"),IF(ABS(E{r})>5,IF(E{r}>0,"🟢 ↑","🟡 ↓"),"⚪ →")))'

        for col in range(1, 8):
            if not is_total:
                ws.cell(row=r, column=col).border = border_thin
            ws.cell(row=r, column=col).alignment = align_right if col > 1 else align_left

    ws.freeze_panes = "B2"
    ws.sheet_properties.pageSetUpPr = None
    return ws


def create_top10(wb):
    ws = wb.create_sheet("Top 10 Variaciones")
    headers = ["Rank", "Partida", "Var. Absoluta", "Var. %", "Tipo Driver", "Clasificación", "Impacto", "Lectura Gerencial"]
    set_col_widths(ws, [8, 30, 18, 12, 18, 18, 15, 40])
    write_header_row(ws, 1, headers)

    for i in range(1, 11):
        r = i + 1
        ws.cell(row=r, column=1, value=i).font = font_data
        ws.cell(row=r, column=1).alignment = align_center
        for col in range(2, 9):
            ws.cell(row=r, column=col).font = font_input
            ws.cell(row=r, column=col).border = border_thin
        ws.cell(row=r, column=8).alignment = align_wrap
        if i % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = fill_gray_light

    ws.freeze_panes = "A2"
    return ws


def create_indicadores(wb, params):
    ws = wb.create_sheet("Indicadores")
    p_base = params.get("periodo_base", "Período Base")
    p_comp = params.get("periodo_comp", "Período Comp.")

    headers = ["Indicador", "Fórmula", p_base, p_comp, "Delta", "Tendencia"]
    set_col_widths(ws, [30, 35, 15, 15, 15, 12])
    write_header_row(ws, 1, headers)

    sections = {
        "MÁRGENES": [
            ("Margen Bruto (Gross Margin)", "Utilidad Bruta / Ingresos × 100"),
            ("Margen Operacional (Operating Margin)", "EBIT / Ingresos × 100"),
            ("Margen EBITDA", "EBITDA / Ingresos × 100"),
            ("Margen Neto (Net Margin)", "Utilidad Neta / Ingresos × 100"),
        ],
        "LIQUIDEZ": [
            ("Razón Corriente (Current Ratio)", "Activos Corrientes / Pasivos Corrientes"),
            ("Prueba Ácida (Quick Ratio)", "(Act. Corrientes - Inventarios) / Pas. Corrientes"),
            ("Capital de Trabajo Neto", "Activos Corrientes - Pasivos Corrientes"),
        ],
        "ENDEUDAMIENTO": [
            ("Endeudamiento Total", "Pasivos / Activos × 100"),
            ("Deuda / Patrimonio (D/E)", "Pasivos / Patrimonio"),
            ("Deuda / EBITDA", "Deuda Financiera / EBITDA"),
        ],
        "RENTABILIDAD": [
            ("ROA (Return on Assets)", "Utilidad Neta / Activos × 100"),
            ("ROE (Return on Equity)", "Utilidad Neta / Patrimonio × 100"),
        ],
        "COBERTURA": [
            ("Cobertura de Intereses", "EBIT / Gastos Financieros"),
        ],
    }

    row = 2
    for section, indicators in sections.items():
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=section)
        c.font = font_section
        c.fill = fill_gray_light
        row += 1
        for name, formula in indicators:
            ws.cell(row=row, column=1, value=name).font = font_data
            ws.cell(row=row, column=2, value=formula).font = font_note
            ws.cell(row=row, column=3).font = font_input
            ws.cell(row=row, column=3).number_format = '0.0'
            ws.cell(row=row, column=4).font = font_input
            ws.cell(row=row, column=4).number_format = '0.0'
            ws.cell(row=row, column=5).value = f'=IF(C{row}="","",D{row}-C{row})'
            ws.cell(row=row, column=5).number_format = '0.0;(0.0);"-"'
            ws.cell(row=row, column=6).value = f'=IF(E{row}="","",IF(E{row}>0,"↑",IF(E{row}<0,"↓","→")))'
            ws.cell(row=row, column=6).alignment = align_center
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border_thin
            row += 1
        row += 1

    ws.freeze_panes = "A2"
    return ws


def create_alertas_recomendaciones(wb):
    ws = wb.create_sheet("Alertas y Recomendaciones")
    set_col_widths(ws, [12, 25, 45, 15, 15])

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="ALERTAS").font = font_section
    ws["A1"].fill = fill_gray_light

    alert_headers = ["Tipo", "Partida/Indicador", "Descripción", "Severidad"]
    write_header_row(ws, 2, alert_headers)
    for i in range(3, 13):
        for col in range(1, 5):
            ws.cell(row=i, column=col).font = font_input
            ws.cell(row=i, column=col).border = border_thin

    rec_row = 15
    ws.merge_cells(f"A{rec_row}:E{rec_row}")
    ws.cell(row=rec_row, column=1, value="RECOMENDACIONES ESTRATÉGICAS").font = font_section
    ws[f"A{rec_row}"].fill = fill_gray_light

    rec_headers = ["#", "Recomendación", "Prioridad", "Horizonte", "Hallazgo Base"]
    write_header_row(ws, rec_row + 1, rec_headers)
    for i in range(rec_row + 2, rec_row + 7):
        ws.cell(row=i, column=1, value=i - rec_row - 1).font = font_data
        ws.cell(row=i, column=1).alignment = align_center
        for col in range(1, 6):
            ws.cell(row=i, column=col).font = font_input
            ws.cell(row=i, column=col).border = border_thin
        ws.cell(row=i, column=2).alignment = align_wrap
        ws.cell(row=i, column=5).alignment = align_wrap

    return ws


def create_metodologia(wb, params):
    ws = wb.create_sheet("Metodología")
    set_col_widths(ws, [80])

    texts = [
        ("NOTAS METODOLÓGICAS", font_section),
        ("", font_data),
        ("Fórmulas utilizadas:", font_label),
        ("• Variación Absoluta = Período Comparación - Período Base", font_data),
        ("• Variación Relativa (%) = (Var. Absoluta / |Período Base|) × 100", font_data),
        ("• Participación = Partida / Total del estado × 100", font_data),
        ("• Índice de Materialidad = |Var. Absoluta| × (Participación / 100)", font_data),
        ("", font_data),
        ("Supuestos aplicados:", font_label),
        ("• Los estados financieros de ambos períodos están expresados bajo las mismas normas contables", font_data),
        ("• La moneda y escala son consistentes entre períodos", font_data),
        ("• Las reclasificaciones contables han sido identificadas y documentadas", font_data),
        ("", font_data),
        ("Criterios de materialidad:", font_label),
        ("• Variación > ±5%: Incluida en análisis de drivers", font_data),
        ("• Participación > 1% del total: Analizada independientemente", font_data),
        ("• Top 10 variaciones absolutas: Siempre incluidas en ranking", font_data),
        ("", font_data),
        ("Señalización:", font_label),
        ("• 🟢 ↑↑ Fuerte mejora: Variación favorable > +15%", font_data),
        ("• 🟢 ↑ Mejora: Variación favorable +5% a +15%", font_data),
        ("• ⚪ → Estable: Variación -5% a +5%", font_data),
        ("• 🟡 ↓ Deterioro: Variación desfavorable -5% a -15%", font_data),
        ("• 🔴 ↓↓ Alerta: Variación desfavorable < -15%", font_data),
        ("• 🔵 Nueva: Partida nueva (base = 0)", font_data),
        ("", font_data),
        ("Fuente de datos:", font_label),
        (f"• Estados financieros proporcionados por {params.get('empresa', '[empresa]')}", font_data),
        ("• Análisis generado con asistencia de IA", font_data),
        ("", font_data),
        ("Limitaciones:", font_label),
        ("• El análisis se basa exclusivamente en la información proporcionada", font_data),
        ("• No se ha realizado auditoría independiente de las cifras", font_data),
        ("• Los drivers identificados son hipótesis basadas en la información disponible", font_data),
        ("• Los indicadores se calculan con los datos disponibles; algunos pueden requerir información adicional", font_data),
        ("", font_data),
        (f"Fecha de elaboración: {datetime.now().strftime('%Y-%m-%d')}", font_note),
        ("Versión: 1.0", font_note),
    ]

    for i, (text, font) in enumerate(texts, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = align_wrap

    return ws


def main():
    parser = argparse.ArgumentParser(description="Genera plantilla Excel para análisis horizontal financiero")
    parser.add_argument("output", help="Ruta del archivo Excel de salida")
    parser.add_argument("--empresa", default="[Nombre de la empresa]")
    parser.add_argument("--tipo", default="[Tipo de EEFF]")
    parser.add_argument("--periodo-base", default="[Período base]", dest="periodo_base")
    parser.add_argument("--periodo-comp", default="[Período comparación]", dest="periodo_comp")
    parser.add_argument("--moneda", default="[Moneda]")
    parser.add_argument("--escala", default="[Escala]")
    args = parser.parse_args()

    params = {
        "empresa": args.empresa,
        "tipo": args.tipo,
        "periodo_base": args.periodo_base,
        "periodo_comp": args.periodo_comp,
        "moneda": args.moneda,
        "escala": args.escala,
    }

    wb = Workbook()
    create_resumen_ejecutivo(wb, params)
    create_analisis_horizontal(wb, params)
    create_top10(wb)
    create_indicadores(wb, params)
    create_alertas_recomendaciones(wb)
    create_metodologia(wb, params)

    wb.save(args.output)
    print(f"Plantilla generada: {args.output}")


if __name__ == "__main__":
    main()
